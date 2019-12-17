# import the necessary modules
import requests
import json
import random
from io import BytesIO
from PIL import Image
import openpyxl
from openpyxl import Workbook
import datetime

#  Part 2: User Picks Art
try:
    # downloading departments data
    response = requests.get('https://collectionapi.metmuseum.org/public/collection/v1/departments')
    # translate test data in JSON format into a dictionary
    dict_departments = json.loads(response.text)
except:
    # error message
    print("Error connect!!!")
    print("Try again!")
    # exit the program
    quit()
# print a list of departments
print("\n The Department names")
for item in dict_departments["departments"]:
    # print ID of a department and the name of a department
    print("{}. {}".format(item["departmentId"], item["displayName"]))

while True:
    # input id of a department
    try:
        ID = int(input("\nInput the number of a department: "))
        # check for the entered ID
        if ID in [item["departmentId"] for item in dict_departments["departments"]]:
            departmentId = ID
            for item in dict_departments["departments"]:
                if item["departmentId"] == departmentId:
                    displayName = item["displayName"]
                    break
            print("\nYou choose {}".format(displayName))
            break
    except :
        # repeat the input in case of incorrect input
        print("Input error!!!")
        print("Try again!")

try:
    # downloading the department data
    response = requests.get('https://collectionapi.metmuseum.org/public/collection/v1/objects?departmentIds={}'.format(departmentId))
    # translate test data in JSON format into a dictionary
    dict_objects = json.loads(response.text)
    # generate a random number between 1 and the maximum number of objects in that department
    number = random.randint(1, dict_objects["total"])
    # define an object
    objectID = dict_objects["objectIDs"][number-1]
    # downloading the object data
    response = requests.get('https://collectionapi.metmuseum.org/public/collection/v1/objects/'+str(objectID))
    # translate test data in JSON format into a dictionary
    art_object = json.loads(response.text)
except:
    # error message
    print("Error connect!!!")
    print("Try again!")
    # exit the program
    quit()

# check that the object is publicly accessible
if art_object["isPublicDomain"]:
    # downloading the small image
    response = requests.get(art_object["primaryImageSmall"])
    # create an image object
    img = Image.open(BytesIO(response.content))
    # show the image
    img.show()
    save = input("Want to save this object?(y/n): ")
    if save == "y":
        # downloading the image
        p = requests.get(art_object["primaryImage"])
        # save this image
        out = open(art_object["objectName"]+".jpg", 'wb')
        out.write(p.content)
        out.close()
    else:
        # exit the program
        quit()
else:
    print("The object is NOT in the public domain.")
    print("Try again!")
    # exit the program
    quit()

# Part 3: Apply a Sticker
# downloading the image
sticker_image = Image.open("monstre.png")

try:
    # create an image object
    art_image = Image.open(art_object["objectName"]+".jpg")
except FileNotFoundError:
    print("File Not Found")
# determine the random coordinates for the sticker placement
x = random.randint(0, art_image.size[0] - sticker_image.size[0])
y = random.randint(0, art_image.size[1] - sticker_image.size[1])
try:
    # determine the angle of rotation
    angle = random.randint(1, 4)
    # rotate by the appropriate number of degrees
    if angle == 1:
        sticker_image = sticker_image.transpose(Image.ROTATE_90)
    if angle == 2:
        sticker_image = sticker_image.transpose(Image.ROTATE_180)
    if angle == 3:
        sticker_image = sticker_image.transpose(Image.ROTATE_270)

    # do or do not mirror
    flip = random.randint(1, 2)
    if flip == 1:
        sticker_image = sticker_image.transpose(Image.FLIP_LEFT_RIGHT)
    # put one image on another
    art_image.paste(sticker_image.convert('RGBA'), (x, y), mask=sticker_image)
except:
    print("The sticker does not fit in the art!")
    print("Try again!")
    # exit the program
    quit()

art_image.show()

# Part 4: Save the Stickered Art
# save the stickered art
art_image.save(art_object["objectName"]+"_monstre.jpg",'JPEG')

# Part 5: Add an Entry to the Spreadsheet
# create a book object
book = Workbook()

sheet = book.active
# define the last filled row of the table
number_rows = sheet.max_row
# write the data into a table
sheet.cell(row=number_rows+1, column=1).value = art_object["title"]
sheet.cell(row=number_rows+1, column=2).value = art_object["artistDisplayName"]
sheet.cell(row=number_rows+1, column=3).value = art_object["primaryImage"]
sheet.cell(row=number_rows+1, column=4).value = "Monstre"
sheet.cell(row=number_rows+1, column=5).value = datetime.datetime.today()
# save the book
book.save('Data.xlsx')



